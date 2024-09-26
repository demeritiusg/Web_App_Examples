import os
import re
import tempfile
from subprocess import Popen
from collections import OrderedDict

import xlrd
import xlwt
import xlutils
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font


"""
Someone was having fun building excel files with this... probably useless
"""

__all__ = ['WorksheetWriter', 'Excel', 'NewExcel', 'ExistingExcel', 'TempExcel',
           'to_excel', 'to_excel_temp', 'to_xl', 'to_xl_tmp']

EXCEL = r'EXCEL.EXE'
EXCEL2003_32 = r'C:\Program Files (x86)\Microsoft Office\Office11\EXCEL.EXE'
EXCEL2007_32 = r'C:\Program Files (x86)\Microsoft Office\Office12\EXCEL.EXE'
EXCEL2007_64 = r'C:\Program Files\Microsoft Office\Office12\EXCEL.EXE'
EXCEL2010_32 = r'C:\Program Files (x86)\Microsoft Office\Office14\EXCEL.EXE'
EXCEL2010_64 = r'C:\Program Files\Microsoft Office\Office14\EXCEL.EXE'
EXCEL2016_64 = r'C:\Program Files\Microsoft Office\root\Office16\EXCEL.EXE'

EXCEL_PATHS = [EXCEL2016_64, EXCEL2010_64, EXCEL2010_32, EXCEL2007_64,
               EXCEL2007_32, EXCEL2003_32, EXCEL]

DEFAULT_DOC_TYPE = '.xlsx'

MAX_SHEETNAME_LENGTH = 31


class WorksheetWriter(object):
    """Writes tabular data by row or by table to a worksheet. Additional sheets
    are automatically created for the XLS format since there is a 65536 row
    limit."""

    def __init__(self, workbook, sheet):
        self.workbook = workbook
        self.sheet = sheet
        self.current_row = 0
        self.overflow_idx = 2
        self.num_cols = None
        self.has_header = False

        try:
            # xlwt
            self.name = sheet.name
        except:
            # openpyxl
            self.name = sheet.title

        self._is_xlsx = isinstance(self.sheet, openpyxl.worksheet.Worksheet)

    def _write(self, row, col, value):
        """Provide consistent write interface between xlwt and openpyxl."""
        if self._is_xlsx:
            c = self.sheet.cell(row=row + 1, column=col + 1)
            c.value = value
        else:
            self.sheet.write(row, col, value)

    def writeheader(self, values):
        """Write the header using writerow() and track that a header has been
        written.
        """
        self.has_header = True
        self.writerow(values)

    def writerow(self, values):
        """Write data by rows. For XLS files, a new sheet with a _[#] suffix
        will be created for every 65536 rows.
        """
        if self.num_cols is None:
            self.num_cols = len(values)

        for col, value in enumerate(values):
            try:
                self._write(self.current_row, col, value)
            except ValueError as e:
                if str(e).find('65536') >= 0:
                    new_sheet = '%s_%s' % (self.sheet.name, self.overflow_idx)
                    self.sheet = self.workbook.create_sheet(title=new_sheet)
                    self.current_row = 0
                    self.overflow_idx += 1

                    self._write(self.current_row, col, value)
                else:
                    raise

        self.current_row += 1

    def writerows(self, rows):
        for row in rows:
            self.writerow(row)

    def writelist(self, lst):
        """Convenience function for writing a single column of values."""
        for itm in lst:
            self.writerow([itm])

    def apply_style(self):
        """Bold and center the header, and apply an auto filter."""
        if self._is_xlsx == False:
            raise Exception('Filters only work for XLSX docs.')
        if self.num_cols is None:
            raise Exception('A row must be written first.')

        font = Font(bold=True)
        alignment = Alignment(horizontal='center')

        header_range = "A1:%s1" % get_column_letter(self.num_cols)
        filter_range = "A1:%s%s" % (get_column_letter(self.num_cols),
                                    self.current_row)
        self.sheet.auto_filter.ref = filter_range
        for row in self.sheet[header_range]:
            for c in row:
                c.font = font
                c.alignment = alignment


class Excel(object):
    """A convience wrapper around xlwt and openpyxl writing facilities to create
    a simple interface for writing tabular data to Excel for both the XLS and
    XLSX formats.

    Notes
    -----
    + Sheetnames are mapped as Python friendly names to the Excel object.
    + NewExcel, New, ExistingExcel, and Existing are provided as wrappers
      around Excel instead of passing the use_existing parameter.
    + An Excel installation is searched by looking at default installation
      paths from newer to older versions of Excel and from 64-bit to 32-bit,
      and finally a PATH relative attempt is made.
    + Temp, TempExcel, to_excel_temp, and to_xl_tmp are provided for quickly
      writing tabular data to an Excel temporary file without having to specify
      a filename.
    + For XLSX files, writeheader() will apply an auto filter, and bold and
      center the header row.
    """

    def __init__(self, file_or_path, sheetnames=None, use_existing=False,
                 encoding='utf-8', excel_paths=EXCEL_PATHS,
                 doc_type=DEFAULT_DOC_TYPE):
        """
        Parameters
        ----------
        file_or_path : str or file-like
            The path to an Excel file to use or create, or a stream to write to.
        sheetnames : str or list
            The name of the sheet or a list of sheet names.
        use_existing : bool
            Use an existing file if it already exists. Raises an exception if the
            file already exists and use_existing is False.
        excel_paths : list of paths
            The list of paths to try in order for the location of EXCEL.EXE
        doc_type : str
            The Excel document type to use. '.xls' or '.xlsx'
        """
        self._file_or_path = file_or_path
        self._use_existing = use_existing
        self._encoding = encoding
        self._excel_paths = excel_paths
        self._doc_type = doc_type
        self._pre_existing = False

        self._sheets = []
        self._worksheet_writers = OrderedDict()

        self._open_workbook()

        if isinstance(sheetnames, str):
            sheetnames = [sheetnames[:MAX_SHEETNAME_LENGTH]]
        else:
            if sheetnames:
                sheetnames = [s[:MAX_SHEETNAME_LENGTH] for s in sheetnames]
            else:
                sheetnames = ['Sheet1']

        self.add_sheets(sheetnames)

    def _to_python_friendly(self, s):
        """Convert to Python friendly version of the sheetname.

        + Strip leading and trailing spaces
        + Spaces to single spaces.
        + Hyphens and spaces to underscore.
        + Remove invalid characters.
        + Remove leading characters until a letter is found.
        + Convert to lowercase.
        + Multiple underscores to single underscore.
        """
        # Strip leading and trailing spaces.
        s = s.strip()
        # Spaces to single space.
        s = re.sub(' +', ' ', s)
        # Hyphens and spaces to underscore.
        s = re.sub(r'[\- ]', '_', s)
        # Remove invalid characters.
        s = re.sub('[^0-9a-zA-Z_]', '', s)
        # Remove leading characters until we find a letter.
        s = re.sub('^[^a-zA-Z]+', '', s)
        # Lowercase.
        s = s.lower()
        # Multiple underscores to one.
        s = re.sub('_+', '_', s)
        return s

    def __getitem__(self, index):
        """Return the WorksheetWriters in the order they were added to the
        workbook.
        """
        if isinstance(index, int):
            worksheet_writer = list(self._worksheet_writers.values())[index]
        elif isinstance(index, str):
            idx = index[:MAX_SHEETNAME_LENGTH]
            worksheet_writer = self._worksheet_writers[idx]
        else:
            raise ValueError("%s is not a valid index." % index)

        return worksheet_writer

    def add_sheet(self, sheetname):
        """Add a new sheet and WorksheetWriter, and append the Python friendly
        sheetnames to this object.
        """
        if sheetname in self._sheets:
            raise ValueError("Worksheet '%s' already exists")

        self._sheets.append(sheetname)
        sheet = self.workbook.create_sheet(title=sheetname)
        worksheet_writer = WorksheetWriter(self.workbook, sheet)
        self._worksheet_writers[sheetname] = worksheet_writer

        method_name = self._to_python_friendly(sheetname)
        if method_name in dir(self):
            raise ValueError(
                "Worksheet name '%s' collides with an existing method. " \
                "Choose a different worksheet name." % method_name
            )
        else:
            self.__dict__[method_name] = worksheet_writer

    def add_sheets(self, sheetnames):
        """Add new WorksheetWriters for the list of sheetnames."""
        for sheetname in sheetnames:
            self.add_sheet(sheetname)

    @property
    def has_header(self):
        return self[0].has_header

    @has_header.setter
    def has_header(self, x):
        self[0].has_header = x

    def writeheader(self, values):
        """Convenience wrapper for calling writeheader() on the first sheet,
        e.g.:
        >>> workbook[0].writeheader([1,2,3])
        """
        self[0].writeheader(values)

    def writerow(self, values):
        """Convenience wrapper for calling writerow() on the first sheet, e.g.:
        >>> workbook[0].writerow([1,2,3])
        """
        self[0].writerow(values)

    def writerows(self, rows):
        """Convenience wrapper for calling writerows() on the first sheet, e.g.:
        >>> workbook[0].writerows([[1,2,3],[4,5,6]])
        """
        self[0].writerows(rows)

    def writelist(self, lst):
        """Convenience wrapper for calling writerows() on the first sheet, e.g.:
        >>> workbook[0].writelist([1,2,3])
        """
        self[0].writelist(lst)

    @property
    def _use_xlsx(self):
        return self._doc_type == '.xlsx'

    def _xl_open_workbook(self):
        """Opens a workbook using either xlrd or openpyxl, and monkey patches
        the resulting workbook objects to have the same interface for our
        purposes.
        """
        if self._use_xlsx:
            wb = openpyxl.load_workbook(self._file_or_path)
            wb.__dict__['sheet_by_name'] = wb.get_sheet_by_name
        else:
            wb = xlutils.copy(xlrd.open_workbook(self._file_or_path))
            wb.__dict__['create_sheet'] = lambda title: wb.add_sheet(title)
        return wb

    def _xl_new_workbook(self):
        """Opens a new workbook using either xlrd or openpyxl, and monkey
        patches the resulting workbook objects to have the same interface for
        our purposes.
        """
        if self._use_xlsx:
            wb = openpyxl.Workbook()  # encoding=self._encoding)
            wb.__dict__['sheet_by_name'] = wb.get_sheet_by_name
            wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
        else:
            wb = xlwt.Workbook(encoding=self._encoding)
            wb.__dict__['create_sheet'] = lambda title: wb.add_sheet(title)
        return wb

    def _open_workbook(self):
        """Return either a xlwt or openpyxl Workbook object that has been monkey
        patched to have a consistent interface for the necessary methods.
        """
        if isinstance(self._file_or_path, str):
            if os.path.exists(self._file_or_path):
                if self._use_existing:
                    workbook = self._xl_open_workbook()
                    for sheetname in self._sheets:
                        try:
                            workbook.sheet_by_name(sheetname)
                        except:
                            pass
                        else:
                            raise IOError("Worksheet '%s' already exists." % sheetname)
                else:
                    raise IOError("The file '%s' already exists." % self._file_or_path)
            else:
                workbook = self._xl_new_workbook()
        else:
            workbook = self._xl_new_workbook()

        self.workbook = workbook

    def save(self):
        """Save the workbook. Also applies the final style to XLSX files."""
        if self._use_xlsx:
            for sheet in self._sheets:
                worksheet_writer = self._worksheet_writers[sheet]
                if worksheet_writer.has_header:
                    worksheet_writer.apply_style()

        if isinstance(self._file_or_path, str):
            if os.path.exists(self._file_or_path):
                os.remove(self._file_or_path)
            self.workbook.save(self._file_or_path)
        else:
            self.workbook.save(self._file_or_path)

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        if exc_type == None:
            self.save()


class New(Excel):
    """A convenience wrapper for Excel(..., use_existing=False)."""

    def __init__(self, file_or_path, sheetname=None):
        super(New, self).__init__(file_or_path, sheetname, use_existing=False)


NewExcel = New


class Existing(Excel):
    """A convenience wrapper for Excel(..., use_existing=True)."""

    def __init__(self, file_or_path, sheetname=None):
        super(Existing, self).__init__(file_or_path, sheetname,
                                       use_existing=True)


ExistingExcel = Existing


class Temp(Excel):
    """Passes a temp file to Excel() and spawns the resulting workbook in the
    EXCEL.EXE application.
    """

    def __init__(self, delete=True, doc_type=DEFAULT_DOC_TYPE):
        self.delete = delete
        self.doc_type = doc_type
        self.tempfile = \
            tempfile.NamedTemporaryFile(suffix=self.doc_type, delete=False)
        super(Temp, self).__init__(self.tempfile.file, doc_type=self.doc_type)

    def __exit__(self, exc_type, exc_value, traceback):
        super(Temp, self).__exit__(exc_type, exc_value, traceback)
        tempfile_path = os.path.abspath(self.tempfile.name)
        self.tempfile.close()

        p = None
        for path in self._excel_paths:
            try:
                p = Popen([path, tempfile_path])
            except:
                pass
            else:
                break

        if p:
            print('... waiting for Excel to exit ...')
            p.wait()
            if os.path.exists(tempfile_path) and self.delete:
                os.remove(tempfile_path)
        else:
            if os.path.exists(tempfile_path) and self.delete:
                os.remove(tempfile_path)
            raise Exception("EXCEL.EXE could not be found.")


TempExcel = Temp


def to_excel(rows, file_or_path,
             header=None, sheetname=None, use_existing=False):
    """A convenience wrapper around Excel to create a workbook with a function
    rather than with a context manager.

    rows : List[List]
        The 2D table array to create the sheet for.

    file_or_path : file or str
        The file to write to.

    header : bool
        True if the first row of `rows` is a header.

    sheetname : str
        The name that will be used for the Excel sheet.

    use_existing : bool
        If True and the workbook exists, insert the sheet, otherwise raise an error.
    """
    with Excel(file_or_path, sheetname, use_existing) as wb:
        idx = sheetname or 0
        if header is True:
            wb[idx].has_header = header
        elif header:
            wb[idx].writeheader(header)
        wb[idx].writerows(rows)


to_xl = to_excel


def to_excel_temp(rows, header=None):
    """A convenience wrapper around Temp to create a workbook with a function
    rather than with a context manager. Creates a temporary file and opens the
    resulting file in the EXCEL.EXE application.

    rows : List[List]
        The 2D table array to create the sheet for.

    header : bool
        The first row of `rows` is the header.
    """
    with Temp() as wb:
        if header is True:
            wb.has_header = header
        elif header:
            wb.writeheader(header)
        wb.writerows(rows)


to_xl_tmp = to_excel_temp

if __name__ == '__main__':
    to_excel_temp([['A', 'B', 'C'], [1, 2, 3]], True)
